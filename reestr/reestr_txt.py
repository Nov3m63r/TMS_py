import csv
import datetime
import openpyxl
import re

file = 'table.xlsx'

unp = 190960286
mailtype = 8
default_weight = 90
default_cost = 2.24
cost_delta = 0.50
d = datetime.datetime.today()
today = f'{d.day}.{d.month}.{d.year}'
lines = 0
reestr_sum = 0
counter = 0
reestr_sheet = []


wb = openpyxl.load_workbook(file)
sheet = wb.active

for row in range(1, sheet.max_row + 1):
    num = sheet.cell(row, 1).value
    if type(num) == int:
        lines += 1
        short_name = sheet.cell(row, 2).value
        quantity = sheet.cell(row, 3).value
        komu = sheet.cell(row, 4).value
        az_etazh = sheet.cell(row, 5).value
        zipcode = sheet.cell(row, 6).value
        obl = sheet.cell(row, 7).value
        raen = sheet.cell(row, 8).value
        gorod = sheet.cell(row, 9).value
        ul = sheet.cell(row, 10).value
        korp = sheet.cell(row, 11).value
        dom = sheet.cell(row, 12).value
        ofis = sheet.cell(row, 13).value

        if zipcode is None:
            zipcode = 220000

        if quantity is None or quantity == 1:
            weight = default_weight
            cost = default_cost
        else:
            short_name += f' {quantity} экз.'
            if quantity == 2:
                weight = default_weight * 2
                cost = default_cost + cost_delta
            elif quantity == 3:
                weight = default_weight * 3
                cost = default_cost + cost_delta * 2
            elif quantity == 4:
                weight = default_weight * 4
                cost = default_cost + cost_delta * 3
            else:
                print(f'In {sheet.row} quantity of magazines is more than 4! ({quantity})')

        reestr_line = f'{num};{short_name};{komu} {az_etazh};;BY;Беларусь;{zipcode};ОПС;' \
                 f'{obl};{raen};{gorod};{ul};{korp};{dom};{ofis}' \
                 f'0;2;0;2;0;0;{weight};0;{cost};0;0'

        reestr_sheet.append(reestr_line)
        reestr_sum += cost

    if num == '@':
        counter += 1
        firstline = f'{unp};{today};{round(reestr_sum, 2)};{lines};{mailtype};1;;;;;;;;;;;;;;;;;;;;'

        with open(f'reestr{counter}.txt', 'w') as reestr_file:
            reestr_file.writelines(firstline + "\n")
            for line in reestr_sheet:
                result1 = re.sub(r'"', '', line)
                result2 = re.sub(r'None', ' ', result1)
                reestr_file.writelines(line + "\n")

        print(f'Reestr sheet {counter} is ready')

        lines = 0
        reestr_sum = 0
        reestr_sheet = []

print(f'Reestr sheets are ready:\n{sheet.max_row} strings in total, {counter} sheets.')
