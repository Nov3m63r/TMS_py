import openpyxl

file = 'table.xlsx'

wb = openpyxl.load_workbook(file)
sheet = wb.active

data = []
for row in range(1, sheet.max_row + 1):
    # num = sheet.cell(row, 1).value
    short_name = sheet.cell(row, 2).value
    quantity = sheet.cell(row, 3).value
    komu = sheet.cell(row, 4).value
    az_etazh = sheet.cell(row, 5).value
    zipcode = sheet.cell(row, 6).value
    obl = sheet.cell(row, 7).value
    raen = sheet.cell(row, 8).value
    gorod = sheet.cell(row, 9).value
    ul = sheet.cell(row, 10).value
    korp = sheet.cell(row, 11).value
    dom = sheet.cell(row, 12).value
    ofis = sheet.cell(row, 13).value

    address = f'{short_name}'
    if komu is not None:
        address += f'@{komu}'
    if quantity is not None:
        address += f'@{quantity} экз.'
    if ul is not None:
        address += f'@ул. {ul}'
    if dom is not None:
        address += f', д. {dom}'
    if korp is not None:
        address += f'/{korp}'
    if ofis is not None:
        address += f', оф. {ofis}'
    if zipcode is not None:
        address += f'@{zipcode}'
    else:
        address += f'@{220000}'
    if gorod is not None:
        address += f', г. {gorod}'
    if obl is not None:
        address += f', {obl} обл.'
    if raen is not None:
        address += f', {raen} р-н'
    if az_etazh is not None:
        address += f', {az_etazh}'

    data.append(address)

with open('address_labels.txt', 'w') as address_file:
    for line in data:
        address_file.writelines(line + "\n")

print(f'Address labels are ready:\n{sheet.max_row} labels in total.')
